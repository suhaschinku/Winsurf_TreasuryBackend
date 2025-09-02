#!/usr/bin/env python3
"""
Simple Document Processor - Clean implementation linking the three components + database storage

Flow:
1. Read file from Documents folder
2. Generate metadata using dublin_core
3. Use metadata and file content in content_scanning  
4. Generate final report
5. Store results in HANA database

No test/example code - production ready implementation.
"""

import os
import json
import logging
from pathlib import Path
from datetime import datetime
from typing import Dict, Any, Optional
# Import the original classes
from Dublin_Core.dublin_core import EnhancedDocumentMetadataExtractor
from Dublin_Core.content_scanner import EnhancedTreasuryContentScanner
from Dublin_Core.content_scanner import DublinCoreMetadataLoader
from Dublin_Core.content_scanner import integrate_enhanced_content_scanning
from Dublin_Core.content_scanner import export_enhanced_scan_results
import glob
import shutil
from pathlib import Path

# Import the enhanced database storage
from Dublin_Core.dublin_core_hana import store_dublin_core_metadata, update_processing_decision

# Simple content extraction
import PyPDF2
import openpyxl

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


class SimpleDocumentProcessor:
    """
    Simple document processor that links dublin_core -> content_scanning -> report -> database
    """
    
    def __init__(self, documents_folder: str = None, output_folder: str = None):
        """
        Initialize processor with folder paths
        
        Args:
            documents_folder: Path to documents folder (default: ./Documents)
            output_folder: Path to output folder (default: ./output)
        """
        self.documents_folder = documents_folder or "./Documents"
        self.output_folder = output_folder or "./output"
        
        # Create output folder if it doesn't exist
        Path(self.output_folder).mkdir(parents=True, exist_ok=True)
        
        # Initialize extractors
        self.metadata_extractor = EnhancedDocumentMetadataExtractor()
        
        logger.info(f"Initialized SimpleDocumentProcessor")
        logger.info(f"Documents folder: {self.documents_folder}")
        logger.info(f"Output folder: {self.output_folder}")
    
    def extract_content(self, file_path: str) -> str:
        """Extract content from document based on file type"""
        file_ext = Path(file_path).suffix.lower()
        
        try:
            if file_ext == '.pdf':
                return self._extract_pdf_content(file_path)
            elif file_ext in ['.xlsx', '.xls']:
                return self._extract_excel_content(file_path)
            elif file_ext in ['.txt', '.csv']:
                return self._extract_text_content(file_path)
            else:
                return f"Unsupported file type: {file_ext}"
        except Exception as e:
            logger.error(f"Content extraction failed: {str(e)}")
            return f"Content extraction error: {str(e)}"
    
    def _extract_pdf_content(self, file_path: str) -> str:
        """Extract text from PDF"""
        content = ""
        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            for page_num in range(min(len(pdf_reader.pages), 50)):
                page = pdf_reader.pages[page_num]
                content += page.extract_text() + "\n"
        return content.strip()
    
    def _extract_excel_content(self, file_path: str) -> str:
        """Extract text from Excel"""
        content = ""
        workbook = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        
        for sheet_name in workbook.sheetnames[:5]:  # Max 5 sheets
            sheet = workbook[sheet_name]
            content += f"\n=== Sheet: {sheet_name} ===\n"
            
            for row in sheet.iter_rows(max_row=50, values_only=True):
                row_data = [str(cell) if cell is not None else "" for cell in row]
                if any(row_data):
                    content += " | ".join(row_data) + "\n"
        
        workbook.close()
        return content.strip()
    
    def _extract_text_content(self, file_path: str) -> str:
        """Extract content from text/csv files"""
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as file:
            return file.read(50000)  # Max 50k chars
    
    def find_file(self, filename: str) -> Optional[str]:
        """Find file in documents folder"""
        documents_path = Path(self.documents_folder)
        
        # Direct path check
        if Path(filename).is_absolute() and Path(filename).exists():
            return filename
        
        # Search in documents folder
        for file_path in documents_path.rglob(filename):
            if file_path.is_file():
                return str(file_path)
        
        return None
    
    def process_document(self, filename: str, store_in_database: bool = True) -> Dict[str, Any]:
        """
        Main processing method - implements the required flow:
        1. Read file from Documents folder
        2. Generate metadata using dublin_core
        3. Use metadata and content in content_scanning
        4. Generate report
        5. Store results in HANA database
        
        Args:
            filename: Name of file to process
            store_in_database: Whether to store results in database (default: True)
            
        Returns:
            Dict with processing results
        """
        logger.info(f"Processing document: {filename}")
        
        result = {
            "filename": filename,
            "timestamp": datetime.now().isoformat(),
            "status": "PROCESSING",
            "file_path": None,
            "metadata_generated": False,
            "content_extracted": False,
            "scanning_completed": False,
            "report_generated": False,
            "database_stored": False,
            "final_decision": None,
            "output_files": [],
            "errors": []
        }
        
        try:
            # Step 1: Find file in Documents folder
            file_path = self.find_file(filename)
            if not file_path:
                result["status"] = "FAILED"
                result["errors"].append(f"File '{filename}' not found in {self.documents_folder}")
                return result
            
            result["file_path"] = file_path
            logger.info(f"Found file: {file_path}")
            
            # Step 2: Generate Dublin Core metadata
            logger.info("Generating Dublin Core metadata...")
            metadata_result = self.metadata_extractor.process_file(file_path, self.output_folder)
            
            if not metadata_result:
                result["status"] = "FAILED"
                result["errors"].append("Dublin Core metadata generation failed")
                return result
            
            result["metadata_generated"] = True
            result["output_files"].extend([
                metadata_result["json_output"],
                metadata_result["ttl_output"]
            ])
            logger.info(f"Metadata generated: {metadata_result['json_output']}")
            
            # Step 3: Extract document content
            logger.info("Extracting document content...")
            document_content = self.extract_content(file_path)
            result["content_extracted"] = True
            logger.info(f"Content extracted: {len(document_content)} characters")
            
            # Step 4: Run content scanning with metadata and content
            logger.info("Running enhanced content scanning...")
            should_proceed, scan_result = integrate_enhanced_content_scanning(
                document_text=document_content,
                filename=filename,
                metadata_dict=metadata_result["metadata"]
            )
            
            result["scanning_completed"] = True
            result["final_decision"] = scan_result.processing_decision.value
            result["should_proceed"] = should_proceed
            result["risk_score"] = scan_result.risk_score
            result["treasury_relevant"] = scan_result.is_treasury_relevant
            result["scb_document"] = scan_result.is_scb_document
            result["reasoning"] = scan_result.reasoning
            result["classification_keywords"] = scan_result.classification_keywords
            result["primary_classification"] = scan_result.primary_classification
            
            # Step 5: Export detailed scan results
            logger.info("Generating final report...")
            scan_results_file = os.path.join(
                self.output_folder, 
                f"{Path(filename).stem}_scan_results.json"
            )
            export_enhanced_scan_results(scan_result, scan_results_file)
            result["output_files"].append(scan_results_file)
            
            # Create summary report
            summary_report = self._create_summary_report(result, scan_result, metadata_result["metadata"])
            summary_file = os.path.join(
                self.output_folder,
                f"{Path(filename).stem}_summary_report.json"
            )
            
            with open(summary_file, 'w', encoding='utf-8') as f:
                json.dump(summary_report, f, indent=2, ensure_ascii=False)
            
            result["output_files"].append(summary_file)
            result["report_generated"] = True
            
            # Step 6: Store in HANA database (if enabled)
            if store_in_database:
                try:
                    logger.info("Storing metadata and processing decision in HANA database...")
                    store_dublin_core_metadata(
                        json_data=metadata_result["metadata"],
                        original_filename=filename,
                        processing_decision=result["final_decision"],
                        reasoning=result["reasoning"],
                        is_scb_document = result["scb_document"],
                        classification_keywords = result["classification_keywords"],
                        primary_classification = result["primary_classification"]
                    )
                    result["database_stored"] = True
                    logger.info("Successfully stored in HANA database")
                except Exception as db_error:
                    logger.error(f"Database storage failed: {str(db_error)}")
                    result["errors"].append(f"Database storage error: {str(db_error)}")
                    # Continue processing even if database fails
            
            result["status"] = "SUCCESS"
            
            # Print summary
            self._print_summary(summary_report)
            
            logger.info(f"Document processing completed successfully: {filename}")
            logger.info(f"Final decision: {result['final_decision']}")
            logger.info(f"Output files: {len(result['output_files'])} files generated")
            if store_in_database:
                logger.info(f"Database storage: {'✅ SUCCESS' if result['database_stored'] else '❌ FAILED'}")
            
        except Exception as e:
            logger.error(f"Processing failed for {filename}: {str(e)}")
            result["status"] = "FAILED"
            result["errors"].append(f"Processing error: {str(e)}")
        
        return result
    
    def update_processing_decision_only(self, filename: str, processing_decision: str, reasoning: str,is_scb_document: bool,
            classification_keywords: str,  primary_classification : str ) -> bool:
        """
        Update only the processing decision and reasoning for an existing record in the database.
        
        Args:
            filename: Original filename
            processing_decision: Final processing decision (APPROVED/NEEDS_APPROVAL/REJECTED)
            reasoning: Reasoning for the processing decision
            
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            logger.info(f"Updating processing decision for {filename}: {processing_decision}")
            update_processing_decision(filename, processing_decision, reasoning,is_scb_document,classification_keywords,primary_classification)
            logger.info(f"Successfully updated processing decision for {filename}")
            return True
        except Exception as e:
            logger.error(f"Failed to update processing decision for {filename}: {str(e)}")
            return False
    
    def _create_summary_report(self, result: Dict[str, Any], scan_result, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """Create concise summary report"""
        return {
            "document_summary": {
                "filename": result["filename"],
                "file_path": result["file_path"],
                "processing_timestamp": result["timestamp"],
                "processing_status": result["status"],
                "database_stored": result["database_stored"]
            },
            "processing_decision": {
                "final_decision": result["final_decision"],
                "should_proceed": result["should_proceed"],
                "risk_score": result["risk_score"],
                "reasoning": scan_result.reasoning
            },
            "document_analysis": {
                "is_treasury_relevant": result["treasury_relevant"],
                "is_scb_document": result["scb_document"],
                "is_personal_document": scan_result.is_personal_document,
                "document_classification": scan_result.document_classification.value,
                "copyright_status": scan_result.copyright_status.value
            },
            "metadata_info": {
                "title": metadata["dublin_core"]["title"],
                "creator": metadata["dublin_core"]["creator"],
                "access_level": metadata["access_control"]["access_level"],
                "file_format": metadata["dublin_core"]["format"],
                "document_type": metadata["dublin_core"]["type"]
            },
            "compliance_info": {
                "prohibited_pii_found": scan_result.prohibited_pii_found,
                "copyright_concerns": scan_result.copyright_concerns,
                "required_approvals": scan_result.required_approvals,
                "recommendations": scan_result.recommendations
            },
            "output_files": result["output_files"],
            "processing_steps": {
                "metadata_generated": result["metadata_generated"],
                "content_extracted": result["content_extracted"],
                "scanning_completed": result["scanning_completed"],
                "report_generated": result["report_generated"],
                "database_stored": result["database_stored"]
            }
        }
    
    def _print_summary(self, summary: Dict[str, Any]) -> None:
        """Print processing summary"""
        print(f"\n{'='*60}")
        print(f"📄 DOCUMENT PROCESSING SUMMARY")
        print(f"{'='*60}")
        
        doc_info = summary["document_summary"]
        decision_info = summary["processing_decision"]
        analysis_info = summary["document_analysis"]
        
        # Status
        status_emoji = "✅" if doc_info["processing_status"] == "SUCCESS" else "❌"
        print(f"{status_emoji} Status: {doc_info['processing_status']}")
        print(f"📁 File: {doc_info['filename']}")
        print(f"💾 Database: {'✅ STORED' if doc_info['database_stored'] else '❌ NOT STORED'}")
        
        # Decision
        decision_emoji = {
            "APPROVED": "✅",
            "NEEDS_APPROVAL": "⚠️", 
            "REJECTED": "🚫"
        }
        
        print(f"\n🎯 Processing Decision: {decision_emoji.get(decision_info['final_decision'], '❓')} {decision_info['final_decision']}")
        print(f"🚀 Should Proceed: {'✅ YES' if decision_info['should_proceed'] else '❌ NO'}")
        print(f"📊 Risk Score: {decision_info['risk_score']:.1f}/100")
        
        # Analysis
        print(f"\n📋 Document Analysis:")
        print(f"   • Treasury Relevant: {'✅ YES' if analysis_info['is_treasury_relevant'] else '❌ NO'}")
        print(f"   • SCB Document: {'✅ YES' if analysis_info['is_scb_document'] else '❌ NO'}")
        print(f"   • Personal Document: {'⚠️ YES' if analysis_info['is_personal_document'] else '✅ NO'}")
        print(f"   • Classification: {analysis_info['document_classification']}")
        
        # Output files
        print(f"\n📁 Generated Files:")
        for file_path in summary["output_files"]:
            print(f"   • {Path(file_path).name}")
        
        print(f"{'='*60}\n")


def cleanup_processing_files(filename: str, output_folder: str = "./output", documents_folder: str = "./Documents"):
    """
    Clean up all intermediary files and source file for a processed document
    Only removes source file if processing decision is not "APPROVED"
    
    Args:
        filename (str): The filename that was processed
        output_folder (str): Path to output folder (default: ./output)
        documents_folder (str): Path to documents folder (default: ./Documents)
    """
    
    cleanup_count = 0
    errors = []
    
    try:
        # Get base filename without extension
        base_filename = Path(filename).stem
        
        # Define specific intermediary files to remove from output folder
        intermediary_files = [
            f"{base_filename}_metadata.json",
            f"{base_filename}_metadata.ttl", 
            f"{base_filename}_scan_results.json",
            f"{base_filename}_summary_report.json"
        ]
        
        # Check processing decision from summary report
        summary_report_path = os.path.join(output_folder, f"{base_filename}_summary_report.json")
        processing_decision = None
        
        if os.path.exists(summary_report_path):
            try:
                with open(summary_report_path, 'r', encoding='utf-8') as f:
                    summary_data = json.load(f)
                    processing_decision = summary_data.get("processing_decision", {}).get("final_decision")
            except Exception as e:
                errors.append(f"Failed to read summary report: {str(e)}")
        
        # Remove intermediary files from output folder
        for file_to_remove in intermediary_files:
            file_path = os.path.join(output_folder, file_to_remove)
            if os.path.exists(file_path):
                try:
                    os.remove(file_path)
                    cleanup_count += 1
                    logger.info(f"Removed: {file_path}")
                except Exception as e:
                    errors.append(f"Failed to remove {file_path}: {str(e)}")
        
        # Remove source file from Documents folder only if NOT APPROVED
        if processing_decision != "APPROVED":
            source_file_path = os.path.join(documents_folder, filename)
            if os.path.exists(source_file_path):
                try:
                    os.remove(source_file_path)
                    cleanup_count += 1
                    logger.info(f"Removed source: {source_file_path}")
                except Exception as e:
                    errors.append(f"Failed to remove source {source_file_path}: {str(e)}")
        else:
            logger.info(f"Source file preserved (APPROVED): {filename}")
        
        # Log results
        if cleanup_count > 0:
            logger.info(f"Cleanup completed: {cleanup_count} files removed for {filename}")
        
        if errors:
            for error in errors:
                logger.error(error)
                
    except Exception as e:
        logger.error(f"Cleanup failed for {filename}: {str(e)}")
        errors.append(f"Cleanup error: {str(e)}")
    
    return cleanup_count, errors