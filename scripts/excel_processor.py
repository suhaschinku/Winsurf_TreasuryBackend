import os
import time
from dotenv import load_dotenv
from langchain.text_splitter import RecursiveCharacterTextSplitter
from gen_ai_hub.proxy.langchain.init_models import init_embedding_model
from langchain.docstore.document import Document
from openpyxl import load_workbook
from logger_setup import get_logger
import re
from concurrent.futures import ThreadPoolExecutor  # For parallel processing
from env_config import EMBEDDING_MODEL
from destination_srv import get_destination_service_credentials, generate_token, fetch_destination_details, extract_aicore_credentials
import requests

logger = get_logger()
load_dotenv()

#$$$ SOC: 28.05.25 -- Initialize AIC Credentials --- $$$#
logger.info("====> excel_processor.py -> AIC CREDENTIALS <====")

# Load VCAP_SERVICES from environment
vcap_services = os.environ.get("VCAP_SERVICES")

# Extract destination service credentials
destination_service_credentials = get_destination_service_credentials(vcap_services)

# Generate OAuth token for destination service
try:
    oauth_token = generate_token(
        uri=destination_service_credentials['dest_auth_url'] + "/oauth/token",
        client_id=destination_service_credentials['clientid'],
        client_secret=destination_service_credentials['clientsecret']
    )
except requests.exceptions.HTTPError as e:
    # Handle HTTP 500 error for invalid client secret
    if e.response is not None and e.response.status_code == 500:
        raise Exception("HTTP 500: Check if the client secret is correct.") from e
    else:
        raise

#-------------------------------- READ AIC Configuration -------------------------------------

# variables for AIC credentials
global AIC_CREDENTIALS

# Get AIC details from Dest Services
dest_AIC = "GENAI_AI_CORE"
aicore_details = fetch_destination_details(
    destination_service_credentials['dest_base_url'],
    dest_AIC,
    oauth_token
)

# Extract AIC Details
AIC_CREDENTIALS = extract_aicore_credentials(aicore_details)
logger.info("excel_processor -> AIC Credential", AIC_CREDENTIALS)

#$$$ EOC: 28.05.25 -- Initialize AIC Credentials --- $$$#

def clean_text(text):
    """Clean OCR artifacts and normalize text."""
    if not text:
        return ""
    text = re.sub(r"JPHORGAN", "JPMORGAN", text, flags=re.IGNORECASE)
    text = re.sub(r"excape", "except", text, flags=re.IGNORECASE)
    text = re.sub(r"\s+", " ", text).strip()  # Normalize whitespace
    return text

def process_single_excel(file_path, filename):
    """Process a single Excel file and extract text."""
    logger.info(f"Processing Excel file: {filename}")
    excel_docs = []
    try:
        wb = load_workbook(file_path, data_only=True)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            text_content = ""
            for row in ws.rows:
                row_text = " ".join([str(cell.value or "").strip() for cell in row if cell.value is not None])
                if row_text:
                    text_content += row_text + "\n"
            if text_content:
                metadata = {
                    "source_file": filename,
                    "sheet": sheet_name,
                    "type": "excel"
                }
                excel_docs.append(Document(page_content=clean_text(text_content.strip()), metadata=metadata))
    except Exception as e:
        logger.info(f"Error processing {filename}: {e}")
    return excel_docs

def save_extracted_data(excel_docs: list[Document], output_dir: str):
    """Save extracted Excel data as a text file."""
    txt_path = os.path.join(output_dir, "excel_data.txt")
    with open(txt_path, "w", encoding="utf-8") as f:
        for doc in excel_docs:
            f.write(f"Source File: {doc.metadata['source_file']}\n")
            f.write(f"Sheet: {doc.metadata['sheet']}\n")
            f.write(f"Content:\n{doc.page_content}\n")
            f.write("-" * 80 + "\n")
    logger.info(f"Saved extracted Excel data to {txt_path}")

def create_embeddings(docs, model_name=EMBEDDING_MODEL, batch_size=400):
    """Generate embeddings for document chunks in batches."""
    if not docs:
        logger.warning("No documents provided for embedding")
        return [], 1
    logger.info(f"Creating embeddings for {len(docs)} docs with {model_name}")
    try:
        ### SOC: Initialize Embedding Models ###
        from gen_ai_hub.proxy import GenAIHubProxyClient
        logger.info("Excel Processor: AIC, {AIC_CREDENTIALS}")

        proxy_client = GenAIHubProxyClient(
            base_url=AIC_CREDENTIALS['aic_base_url'],
            auth_url=AIC_CREDENTIALS['aic_auth_url'],
            client_id=AIC_CREDENTIALS['clientid'],
            client_secret=AIC_CREDENTIALS['clientsecret'],
            resource_group=AIC_CREDENTIALS['resource_group']
        )

        embedding_model = init_embedding_model(model_name=EMBEDDING_MODEL, proxy_client=proxy_client)
        ### EOC: Initialize Embedding Models ###

        logger.info(f"Embedding model initialized: {embedding_model}")
        results = []
        for i in range(0, len(docs), batch_size):
            batch = docs[i:i + batch_size]
            batch_texts = [doc.page_content for doc in batch]
            if not batch_texts or any(not isinstance(text, str) for text in batch_texts):
                logger.info(f"Invalid batch content at index {i}: {batch_texts[:50]}")
                return [], 1
            logger.info(f"Processing batch {i//batch_size + 1}, size: {len(batch_texts)}, Sample: {batch_texts[0][:50]}")
            embeddings = embedding_model.embed_documents(batch_texts)
            if not embeddings or len(embeddings) != len(batch_texts):
                logger.info(f"Embedding mismatch: {len(embeddings)} embeddings for {len(batch_texts)} texts")
                return [], 1
            results.extend(zip(batch, embeddings))
            logger.info(f"Batch {i//batch_size + 1} completed")
            if i + batch_size < len(docs):
                time.sleep(10)  # Rate limiting delay
        logger.info(f"Created embeddings for {len(results)} docs")
        return results, 0
    except Exception as e:
        logger.info(f"Embedding error: {str(e)}", exc_info=True)
        return [], 1

def process_excel(excel_dir):
    """Extract text from irregular Excel files, returning chunks."""
    logger.info(f"Processing Excel files in: {excel_dir}")
    excel_files = [f for f in os.listdir(excel_dir) if f.lower().endswith(('.xlsx', '.xls'))]
    if not excel_files:
        logger.warning(f"No Excel files found in {excel_dir}")
        return []
    all_excel_docs = []

    def process_single_file(filename):
        """Process a single Excel file and return its chunks."""
        file_path = os.path.join(excel_dir, filename)
        return process_single_excel(file_path, filename)

    # Use ThreadPoolExecutor for parallel processing
    with ThreadPoolExecutor(max_workers=4) as executor:  # Adjust max_workers as needed
        futures = [executor.submit(process_single_file, filename) for filename in excel_files]
        for future in futures:
            try:
                all_excel_docs.extend(future.result())
            except Exception as e:
                logger.info(f"Error processing an Excel file: {e}", exc_info=True)

    if not all_excel_docs:
        logger.warning(f"No content extracted from Excel files in {excel_dir}")
        return []
    # Save extracted data as text file
    #save_extracted_data(all_excel_docs, excel_dir)
    # Split into chunks
    text_splitter = RecursiveCharacterTextSplitter(
        chunk_size=10000,
        chunk_overlap=400,
        length_function=len
    )
    chunks = text_splitter.split_documents(all_excel_docs)
    if not chunks:
        logger.warning(f"No chunks created from Excel files")
        return []
    logger.info(f"Processed {len(chunks)} chunks from Excel files")
    return chunks

def process_all_excel(excel_dir, model=EMBEDDING_MODEL):
    """Process all Excel files in a directory and generate embeddings."""
    if not os.path.exists(excel_dir):
        logger.info(f"Excel directory not found: {excel_dir}")
        print(f"Excel directory not found: {excel_dir}")
        return []
    excel_files = [f for f in os.listdir(excel_dir) if f.lower().endswith(('.xlsx', '.xls'))]
    print(f"Found {len(excel_files)} Excel files")
    if not excel_files:
        print(f"No Excel files in {excel_dir}")
        return []
    # Process Excel files into chunks
    chunks = process_excel(excel_dir)
    if not chunks:
        print(f"No chunks extracted from Excel files in {excel_dir}")
        return []
    # Generate embeddings
    embeddings, error = create_embeddings(chunks, model)
    if error:
        print(f"❌ Embedding creation failed for Excel files!")
    else:
        print(f"✅ Created {len(embeddings)} embeddings for Excel files")
    return embeddings